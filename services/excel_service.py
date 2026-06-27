"""Excel 报告导出服务。"""

from pathlib import Path

from models.sensor_data import SensorData
from models.test_record import TestRecord


def export_excel_report(record: TestRecord, samples: list[SensorData], result: dict) -> Path:
    """导出 Excel 报告。"""
    out_dir = Path("reports")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{record.testid}_报告.xlsx"

    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
    except ImportError:
        out_path = out_path.with_suffix(".txt")
        out_path.write_text("未安装 openpyxl，无法生成 Excel。请执行 pip install -r requirements.txt", encoding="utf-8")
        return out_path

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # ── 样式定义 ──
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    label_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    label_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "试验信息"

    # 标题行
    ws_info.merge_cells("A1:B1")
    ws_info["A1"] = "ISO 11820 建筑材料不燃性试验报告"
    ws_info["A1"].font = Font(bold=True, size=14)
    ws_info["A1"].alignment = Alignment(horizontal="center")

    info_rows = [
        ("样品编号", record.productid),
        ("试验编号", record.testid),
        ("操作员", record.operator),
        ("试验前质量(g)", record.preweight),
        ("试验后质量(g)", result["postweight"]),
        ("失重率(%)", round(result["lostweight_per"], 2)),
        ("样品温升(℃)", round(result["deltatf"], 2)),
    ]
    for i, (label, value) in enumerate(info_rows, start=3):
        cell_l = ws_info.cell(row=i, column=1, value=label)
        cell_l.font = label_font
        cell_l.fill = label_fill
        cell_l.border = thin_border
        cell_v = ws_info.cell(row=i, column=2, value=value)
        cell_v.border = thin_border
        cell_v.alignment = center_align

    # 判定结论
    conclusion_row = 3 + len(info_rows) + 1
    ws_info.merge_cells(f"A{conclusion_row}:B{conclusion_row}")
    conclusion = _judge_iso_conclusion(result)
    cell_c = ws_info.cell(row=conclusion_row, column=1, value=f"判定结论：{conclusion}")
    cell_c.font = Font(bold=True, size=12, color="FF0000" if "不合格" in conclusion else "008000")

    ws_info.column_dimensions["A"].width = 22
    ws_info.column_dimensions["B"].width = 28

    # ── 温度数据（带样式） ──
    ws_data = wb.create_sheet("温度数据")
    headers = ["Time", "Temp1", "Temp2", "TempSurface", "TempCenter", "TempCalibration"]
    for ci, h in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    for ri, sample in enumerate(samples, start=2):
        for ci, val in enumerate([
            sample.time_seconds, sample.tf1, sample.tf2,
            sample.ts, sample.tc, sample.tcal,
        ], start=1):
            cell = ws_data.cell(row=ri, column=ci, value=val)
            cell.border = thin_border

    # ── 温度曲线 ──
    ws_chart = wb.create_sheet("温度曲线")
    chart = LineChart()
    chart.title = "ISO11820 温度曲线"
    chart.y_axis.title = "温度(℃)"
    chart.x_axis.title = "时间(s)"
    if len(samples) >= 2:
        values = Reference(ws_data, min_col=2, max_col=5, min_row=1, max_row=len(samples) + 1)
        cats = Reference(ws_data, min_col=1, min_row=2, max_row=len(samples) + 1)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
        ws_chart.add_chart(chart, "A1")

    wb.save(out_path)
    return out_path


def _judge_iso_conclusion(result: dict) -> str:
    """根据 ISO 11820 简化判定规则输出结论。"""
    deltatf = result.get("deltatf", 0)
    lostweight = result.get("lostweight_per", 0)
    flameduration = result.get("flameduration", 0)

    reasons = []
    if deltatf > 30:
        reasons.append(f"温升 {deltatf:.1f}℃ > 30℃")
    if lostweight > 50:
        reasons.append(f"失重率 {lostweight:.1f}% > 50%")
    if flameduration > 0:
        reasons.append(f"火焰持续 {flameduration}s > 0s")

    if not reasons:
        return "合格（A1 级）"
    return f"不合格 — {'、'.join(reasons)}"
